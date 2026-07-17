"""Builds the categorized transactions + category-subtotal Excel workbook."""
import io

import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

HEADER_FILL = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True)
CURRENCY_FORMAT = '#,##0.00;[Red](#,##0.00)'


def _style_header(ws, ncols, row=1):
    for col in range(1, ncols + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center")


def _autofit(ws, df, extra=2):
    for i, col in enumerate(df.columns, start=1):
        max_len = max([len(str(col))] + [len(str(v)) for v in df[col].astype(str)])
        ws.column_dimensions[get_column_letter(i)].width = min(max_len + extra, 60)


def build_workbook(transactions_df: pd.DataFrame) -> bytes:
    """
    transactions_df must have columns: Date, CheckNumber, Description,
    Category, Type, Amount [, SourceFile]
    Returns the .xlsx file as bytes.
    """
    export_cols = ["Date", "CheckNumber", "Description", "Category", "Type", "Amount"]
    if "SourceFile" in transactions_df.columns:
        export_cols.append("SourceFile")
    tx = transactions_df[export_cols].rename(columns={
        "CheckNumber": "Check Number",
        "SourceFile": "Source File",
    }).sort_values("Date")

    summary = (
        transactions_df.groupby("Category", as_index=False)["Amount"]
        .agg(Total="sum", Count="count")
        .sort_values("Total")
    )
    grand_total = pd.DataFrame([{
        "Category": "Grand Total",
        "Total": transactions_df["Amount"].sum(),
        "Count": len(transactions_df),
    }])
    summary = pd.concat([summary, grand_total], ignore_index=True)

    buffer = io.BytesIO()
    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        summary.to_excel(writer, sheet_name="Summary", index=False)
        tx.to_excel(writer, sheet_name="Transactions", index=False)

        wb = writer.book

        ws_summary = wb["Summary"]
        _style_header(ws_summary, len(summary.columns))
        _autofit(ws_summary, summary)
        for row in range(2, ws_summary.max_row + 1):
            ws_summary.cell(row=row, column=2).number_format = CURRENCY_FORMAT
        ws_summary.cell(row=ws_summary.max_row, column=1).font = Font(bold=True)
        ws_summary.cell(row=ws_summary.max_row, column=2).font = Font(bold=True)
        ws_summary.freeze_panes = "A2"

        ws_tx = wb["Transactions"]
        _style_header(ws_tx, len(tx.columns))
        _autofit(ws_tx, tx)
        amount_col = list(tx.columns).index("Amount") + 1
        date_col = list(tx.columns).index("Date") + 1
        for row in range(2, ws_tx.max_row + 1):
            ws_tx.cell(row=row, column=amount_col).number_format = CURRENCY_FORMAT
            ws_tx.cell(row=row, column=date_col).number_format = "mm/dd/yyyy"
        ws_tx.freeze_panes = "A2"
        ws_tx.auto_filter.ref = ws_tx.dimensions

    buffer.seek(0)
    return buffer.getvalue()
